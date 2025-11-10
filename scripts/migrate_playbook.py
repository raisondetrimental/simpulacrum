"""
Migration script to extract data from The Playbook Excel file
and create JSON data files for web-based management.

Usage:
    python scripts/migrate_playbook.py
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

# Add backend to path for openpyxl
sys.path.insert(0, str(Path(__file__).parent.parent / 'backend'))

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def excel_date_to_iso(value: Any) -> Optional[str]:
    """Convert Excel datetime to ISO format string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        # Try to parse string dates
        try:
            dt = datetime.fromisoformat(value)
            return dt.isoformat()
        except:
            return value
    return str(value)


def safe_str(value: Any) -> str:
    """Safely convert value to string."""
    if value is None:
        return ""
    return str(value).strip()


def safe_float(value: Any) -> Optional[float]:
    """Safely convert value to float."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_int(value: Any) -> Optional[int]:
    """Safely convert value to int."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def safe_bool(value: Any) -> bool:
    """Safely convert value to boolean (for team member checkboxes)."""
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ('yes', 'y', 'true', 'x', '1')
    if isinstance(value, (int, float)):
        return value != 0
    return False


def migrate_external_contacts(ws) -> List[Dict[str, Any]]:
    """Extract External Contacts sheet data."""
    contacts = []

    # Skip header row, start from row 2
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        # Skip empty rows
        if not any(row):
            continue

        contact = {
            "id": f"playbook_contact_{idx:03d}",
            "name": safe_str(row[0]),
            "email": safe_str(row[1]),
            "role": safe_str(row[2]),
            "contact_level": safe_int(row[3]) or 3,  # Default to 3 if null
            "region": safe_str(row[4]),
            "last_contact": excel_date_to_iso(row[5]),
            "should_contact": excel_date_to_iso(row[6]),
            "notes": safe_str(row[7])
        }

        # Only add if has at least a name
        if contact["name"]:
            contacts.append(contact)

    return contacts


def migrate_calendar(ws) -> List[Dict[str, Any]]:
    """Extract Calendar sheet data."""
    calendar_entries = []

    # Skip header row, start from row 2
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        # Skip empty rows
        if not any(row):
            continue

        entry = {
            "id": f"playbook_cal_{idx:03d}",
            "date": excel_date_to_iso(row[0]),
            "tasks": safe_str(row[1]),
            "internal_ents": safe_str(row[2]),
            "external_ents": safe_str(row[3]),
            "where": safe_str(row[4]),
            "other_notes": safe_str(row[5]),
            "other_external": safe_str(row[6]),
            "nav": safe_bool(row[7]) if len(row) > 7 else False,
            "aijan": safe_bool(row[8]) if len(row) > 8 else False,
            "lavinia": safe_bool(row[9]) if len(row) > 9 else False,
            "kush": safe_bool(row[10]) if len(row) > 10 else False,
            "amgalan": safe_bool(row[11]) if len(row) > 11 else False,
            "max": safe_bool(row[12]) if len(row) > 12 else False
        }

        # Only add if has a date or tasks
        if entry["date"] or entry["tasks"]:
            calendar_entries.append(entry)

    return calendar_entries


def migrate_people(ws) -> List[Dict[str, Any]]:
    """Extract People sheet data."""
    people = []

    # Skip header row, start from row 2
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        # Skip empty rows
        if not any(row):
            continue

        person = {
            "id": f"playbook_person_{idx:03d}",
            "team_member": safe_str(row[0]),
            "location": safe_str(row[1]),
            "role": safe_str(row[2]),
            "tasks": safe_str(row[3]),
            "disc_profile": safe_str(row[4]),
            "facts_interests": safe_str(row[5])
        }

        # Only add if has a team member name
        if person["team_member"]:
            people.append(person)

    return people


def migrate_deal_flow(ws) -> List[Dict[str, Any]]:
    """Extract Deal Flow sheet data."""
    deals = []

    # Skip header row, start from row 2
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        # Skip empty rows
        if not any(row):
            continue

        deal = {
            "id": f"playbook_deal_{idx:03d}",
            "mu_id": safe_str(row[0]),
            "deal_acronym": safe_str(row[1]),
            "deal": safe_str(row[2]),
            "fx": safe_str(row[3]),
            "total_facility": safe_float(row[4]),
            "sponsor": safe_str(row[5]),
            "financial_close": excel_date_to_iso(row[6]),
            "lead": safe_str(row[7]),
            "type": safe_str(row[8]),
            "security": safe_str(row[9]),
            "benchmark": safe_str(row[10]),
            "benchmark_value": safe_float(row[11]),
            "spread": safe_float(row[12]),
            "rate": safe_float(row[13]) if len(row) > 13 else None
        }

        # Only add if has at least a deal name or MU ID
        if deal["deal"] or deal["mu_id"]:
            deals.append(deal)

    return deals


def migrate_workstream(ws) -> List[Dict[str, Any]]:
    """Extract Workstream sheet data with hierarchical structure.

    Structure:
    - Rows with value in column A (Mission Goals) are parent tasks
    - Rows with None in column A but other content are subtasks
    - Columns: Mission Goals, Process, Category, Deliverable, Done, _, Key, Description
    """
    workstreams = []
    current_parent = None
    parent_id_counter = 1

    # Skip header row, start from row 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if not any(row):
            continue

        mission_goal = safe_str(row[0]) if len(row) > 0 else ""
        process = safe_str(row[1]) if len(row) > 1 else ""
        category = safe_str(row[2]) if len(row) > 2 else ""
        deliverable = safe_str(row[3]) if len(row) > 3 else ""
        done = safe_str(row[4]) if len(row) > 4 else ""
        key = safe_str(row[6]) if len(row) > 6 else ""
        description = safe_str(row[7]) if len(row) > 7 else ""

        # If column A (Mission Goals) has value, this is a parent task
        if mission_goal:
            # Save previous parent if it exists
            if current_parent is not None:
                workstreams.append(current_parent)

            # Create new parent task
            current_parent = {
                "id": f"playbook_workstream_{parent_id_counter:03d}",
                "mission_goal": mission_goal,
                "process": process,
                "category": category,
                "deliverable": deliverable,
                "done": done.lower() in ('done', 'yes', 'x', 'true') if done else False,
                "key": key,
                "description": description,
                "completed": False,  # Checkbox for completion tracking
                "subtasks": []
            }
            parent_id_counter += 1

        # If column A is empty but other columns have content, this is a subtask
        elif not mission_goal and (process or deliverable or category):
            if current_parent is None:
                # Orphan subtask (shouldn't happen, but handle gracefully)
                continue

            subtask_id = len(current_parent["subtasks"]) + 1
            subtask = {
                "id": f"{current_parent['id']}_sub_{subtask_id}",
                "process": process,
                "category": category,
                "deliverable": deliverable,
                "done": done.lower() in ('done', 'yes', 'x', 'true') if done else False,
                "completed": False  # Checkbox for completion tracking
            }
            current_parent["subtasks"].append(subtask)

    # Don't forget to add the last parent
    if current_parent is not None:
        workstreams.append(current_parent)

    return workstreams


def migrate_filing(ws) -> Dict[str, str]:
    """Extract Filing sheet data (instructional content)."""
    # Collect all text from the sheet
    content_lines = []

    for row in ws.iter_rows(values_only=True):
        row_text = " ".join([safe_str(cell) for cell in row if cell])
        if row_text.strip():
            content_lines.append(row_text.strip())

    return {
        "content": "\n\n".join(content_lines)
    }


def save_json(data: Any, filename: str, data_dir: Path) -> None:
    """Save data to JSON file with timestamped backup."""
    filepath = data_dir / filename

    # Create timestamped backup if file exists
    if filepath.exists():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = filepath.stem + f'_backup_{timestamp}' + filepath.suffix
        backup_path = filepath.parent / backup_name

        # Copy to backup (don't rename, in case something goes wrong)
        import shutil
        shutil.copy2(filepath, backup_path)
        print(f"  Created backup: {backup_name}")

    # Save new data
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"  Created: {filename}")


def main():
    """Main migration function."""
    print("=" * 60)
    print("Playbook Excel to JSON Migration")
    print("=" * 60)

    # Define paths
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / "backend" / "data" / "excel" / "The Playbook - Copy.xlsx"
    json_dir = base_dir / "backend" / "data" / "json"

    # Check if Excel file exists
    if not excel_path.exists():
        print(f"\nError: Excel file not found at: {excel_path}")
        sys.exit(1)

    print(f"\nReading from: {excel_path}")
    print(f"Saving to: {json_dir}\n")

    try:
        # Load workbook
        print("Loading Excel workbook...")
        wb = load_workbook(excel_path, read_only=True, data_only=True)

        # Get sheet names
        print(f"Found {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}\n")

        # Migrate each sheet
        migrations = [
            ("External Contacts", "playbook_contacts.json", migrate_external_contacts),
            ("Calendar", "playbook_calendar.json", migrate_calendar),
            ("People", "playbook_people.json", migrate_people),
            ("Deal Flow", "playbook_deals.json", migrate_deal_flow),
            ("Workstream", "playbook_workstreams.json", migrate_workstream),
            ("Filing", "playbook_filing.json", migrate_filing)
        ]

        for sheet_name, json_file, migrate_func in migrations:
            print(f"Processing '{sheet_name}' sheet...")

            if sheet_name not in wb.sheetnames:
                print(f"  Warning: Sheet '{sheet_name}' not found, skipping...")
                continue

            ws = wb[sheet_name]
            data = migrate_func(ws)

            # Show count for list-based data
            if isinstance(data, list):
                print(f"  Extracted {len(data)} records")

            save_json(data, json_file, json_dir)
            print()

        wb.close()

        print("=" * 60)
        print("Migration completed successfully!")
        print("=" * 60)

    except Exception as e:
        print(f"\nError during migration: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

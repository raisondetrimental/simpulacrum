"""
Quick script to examine the Workstreams sheet structure
"""
import openpyxl
from pathlib import Path

# Path to Excel file
excel_path = Path(__file__).parent.parent / 'backend' / 'data' / 'excel' / 'The Playbook - Copy.xlsx'

# Load workbook
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

# First, list all sheet names
print('Available sheets:')
print(wb.sheetnames)
print()

# Try to find workstreams sheet
workstreams_sheet = None
for sheet_name in wb.sheetnames:
    if 'workstream' in sheet_name.lower() or 'mission' in sheet_name.lower():
        workstreams_sheet = sheet_name
        break

if not workstreams_sheet:
    print('Could not find workstreams sheet. Exiting.')
    wb.close()
    exit()

print(f'Using sheet: {workstreams_sheet}')
print('=' * 80)

ws = wb[workstreams_sheet]

# Get headers
headers = [cell.value for cell in ws[1]]
print('Headers:', headers)
print()

# Print all rows
print('All data rows:')
print('-' * 80)
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if any(cell is not None for cell in row):  # Only print non-empty rows
        print(f'Row {i}: {row}')

wb.close()

#!/usr/bin/env python3
"""
ETL script to read Markets Dashboard Excel file and generate JSON for web application.
Reads the Dashboard sheet in read-only mode (no macro execution) and extracts:
- Sovereign Yields by country and maturity
- Corporate Yields by credit rating
- FX Rates and changes
- Central Bank Policy Rates
- Credit Ratings analysis
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
import pandas as pd


class DashboardETL:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.dashboard_sheet = None
        self.data = {
            'metadata': {
                'last_updated': datetime.now().isoformat(),
                'source_file': os.path.basename(excel_file_path),
                'generated_by': 'read_dashboard.py'
            },
            'sections': {}
        }

    def load_workbook(self):
        """Load the Excel workbook in read-only mode"""
        try:
            self.workbook = load_workbook(self.excel_file_path, read_only=True, data_only=True)
            if 'Dashboard' not in self.workbook.sheetnames:
                raise ValueError("Dashboard sheet not found in workbook")
            self.dashboard_sheet = self.workbook['Dashboard']
            print(f"Loaded workbook: {self.excel_file_path}")
            print(f"Dashboard sheet dimensions: {self.dashboard_sheet.max_row}x{self.dashboard_sheet.max_column}")
        except Exception as e:
            raise RuntimeError(f"Failed to load workbook: {e}")

    def get_cell_value(self, row, col):
        """Get cell value with error handling"""
        try:
            cell = self.dashboard_sheet.cell(row=row, column=col)
            return cell.value
        except:
            return None

    def get_row_data(self, row_num, start_col=1, end_col=None):
        """Get all data from a row"""
        if end_col is None:
            end_col = self.dashboard_sheet.max_column

        row_data = []
        for col in range(start_col, end_col + 1):
            value = self.get_cell_value(row_num, col)
            row_data.append(value)
        return row_data

    def extract_sovereign_yields(self):
        """Extract sovereign yields data (around rows 12-35)"""
        print("Extracting Sovereign Yields...")

        sovereign_data = {
            'domestic_currency': {},
            'usd_denominated': {}
        }

        # Domestic Currency section (around row 15-19)
        countries = []
        for col in range(5, 8):  # Columns E, F, G
            country = self.get_cell_value(15, col)
            if country:
                countries.append(str(country).strip())

        maturities = ['1', '3', '5', '10']

        for i, maturity in enumerate(maturities):
            row_num = 16 + i
            yields = []
            for col in range(5, 5 + len(countries)):
                yield_val = self.get_cell_value(row_num, col)
                if yield_val is not None:
                    yields.append(float(yield_val) * 100 if isinstance(yield_val, (int, float)) and yield_val < 1 else yield_val)
                else:
                    yields.append(None)

            sovereign_data['domestic_currency'][f'{maturity}Y'] = dict(zip(countries, yields))

        # USD Denominated section (around row 22-33)
        # USA rates with changes
        usa_data = {}
        for i, maturity in enumerate(maturities):
            row_num = 23 + i
            rate = self.get_cell_value(row_num, 5)  # Column E
            change_1d = self.get_cell_value(row_num, 6)  # Column F
            change_1w = self.get_cell_value(row_num, 7)  # Column G
            change_1m = self.get_cell_value(row_num, 8)  # Column H

            usa_data[f'{maturity}Y'] = {
                'rate': float(rate) * 100 if isinstance(rate, (int, float)) and rate < 1 else rate,
                'changes': {
                    '1D': change_1d,
                    '1W': change_1w,
                    '1M': change_1m
                }
            }

        sovereign_data['usd_denominated'] = {'USA': usa_data}

        # Other countries USD denominated (around row 29-33)
        other_countries = []
        for col in range(5, 8):
            country = self.get_cell_value(29, col)
            if country:
                other_countries.append(str(country).strip())

        for country in other_countries:
            if country not in sovereign_data['usd_denominated']:
                sovereign_data['usd_denominated'][country] = {}

        for i, maturity in enumerate(maturities):
            row_num = 30 + i
            for j, country in enumerate(other_countries):
                rate = self.get_cell_value(row_num, 5 + j)
                if rate is not None:
                    if country not in sovereign_data['usd_denominated']:
                        sovereign_data['usd_denominated'][country] = {}
                    sovereign_data['usd_denominated'][country][f'{maturity}Y'] = {
                        'rate': float(rate) * 100 if isinstance(rate, (int, float)) and rate < 1 else rate
                    }

        self.data['sections']['sovereign_yields'] = sovereign_data

    def extract_corporate_yields(self):
        """Extract corporate yields data (around rows 36-44)"""
        print("Extracting Corporate Yields...")

        corporate_data = {}
        ratings = ['AAA', 'AA', 'A', 'BBB', 'BB', 'High Yield']

        for i, rating in enumerate(ratings):
            row_num = 39 + i
            rate = self.get_cell_value(row_num, 5)
            change_1d = self.get_cell_value(row_num, 6)
            change_1w = self.get_cell_value(row_num, 7)
            change_1m = self.get_cell_value(row_num, 8)

            corporate_data[rating] = {
                'effective_yield': float(rate) * 100 if isinstance(rate, (int, float)) and rate < 1 else rate,
                'changes': {
                    '1D': change_1d,
                    '1W': change_1w,
                    '1M': change_1m
                }
            }

        self.data['sections']['corporate_yields'] = corporate_data

    # FX rates now fetched from ExchangeRate-API (see backend/src/api/fx_rates.py)
    # def extract_fx_rates(self):
    #     """Extract FX rates data (around rows 47-53) - DEPRECATED"""
    #     print("Extracting FX Rates...")
    #
    #     fx_data = {}
    #
    #     # Start from row 50 where data begins
    #     for row_num in range(50, 54):  # Check a few rows for FX data
    #         currency_name = self.get_cell_value(row_num, 3)
    #         identifier = self.get_cell_value(row_num, 4)
    #         rate = self.get_cell_value(row_num, 5)
    #         change_1d = self.get_cell_value(row_num, 6)
    #         change_1w = self.get_cell_value(row_num, 7)
    #         change_1m = self.get_cell_value(row_num, 8)
    #
    #         if currency_name and identifier:
    #             fx_data[str(identifier).upper()] = {
    #                 'name': str(currency_name),
    #                 'rate': rate,
    #                 'changes': {
    #                     '1D': change_1d if change_1d != '#DIV/0!' else None,
    #                     '1W': change_1w if change_1w != '#DIV/0!' else None,
    #                     '1M': change_1m if change_1m != '#DIV/0!' else None
    #                 }
    #             }
    #
    #     self.data['sections']['fx_rates'] = fx_data

    def extract_central_bank_rates(self):
        """Extract central bank policy rates (around row 90+)"""
        print("Extracting Central Bank Policy Rates...")

        cb_data = {}

        # Central bank data is at rows 92-95 with format: Empty | Empty | Country | "% p.a" | Rate
        for row_num in range(92, 96):
            country = self.get_cell_value(row_num, 3)  # Column C (country name)
            rate = self.get_cell_value(row_num, 5)     # Column E (rate value)

            if country and isinstance(country, str) and rate is not None and isinstance(rate, (int, float)):
                country = str(country).strip()
                if rate > 0:
                    # Convert to percentage if needed
                    if rate < 1:
                        rate_pct = float(rate) * 100
                    else:
                        rate_pct = float(rate)

                    cb_data[country] = {'policy_rate': rate_pct}
                    print(f"  Found: {country} = {rate_pct:.2f}%")

        self.data['sections']['central_bank_rates'] = cb_data

    def extract_credit_ratings(self):
        """Extract credit ratings data from Comparable Sovereign Yields section (rows 220-241)"""
        print("Extracting Credit Ratings...")

        ratings_data = {}

        # Extract from "Comparable Sovereign Yields by Rating" section
        # Format: Rating (Col 3) | "% p.a" (Col 4) | 10Y Bond Yield (Col 6)
        # Data starts at row 220 and goes to row 241
        for row_num in range(220, 242):
            rating = self.get_cell_value(row_num, 3)      # Rating in column 3
            yield_val = self.get_cell_value(row_num, 6)   # Yield in column 6

            if rating and isinstance(rating, str) and yield_val and isinstance(yield_val, (int, float)):
                rating = rating.strip()

                # Convert yield to percentage (values are in decimal form like 0.014 = 1.4%)
                yield_pct = float(yield_val) * 100

                ratings_data[rating] = {
                    'benchmark_yields': {
                        '10Y': yield_pct
                    }
                }
                print(f"  Found rating {rating}: 10Y = {yield_pct:.3f}%")

        self.data['sections']['credit_ratings'] = ratings_data

    def extract_benchmark_yields(self):
        """Extract benchmark yields data"""
        print("Extracting Benchmark Yields...")

        benchmark_data = {}

        # Look around row 55-65 for benchmark data
        for row_num in range(55, 75):
            row_data = self.get_row_data(row_num)

            # Look for rows with yield data
            if any(isinstance(val, (int, float)) and 0 < val < 1 for val in row_data):
                country = None
                maturity = None
                yield_val = None

                # Extract country and maturity info from nearby cells
                for col in range(1, 6):
                    val = self.get_cell_value(row_num, col)
                    if val and isinstance(val, str):
                        if len(val) > 2:
                            country = val.strip()
                        elif val.isdigit():
                            maturity = f"{val}Y"

                # Find the yield value
                for val in row_data:
                    if isinstance(val, (int, float)) and 0 < val < 1:
                        yield_val = float(val) * 100
                        break

                if country and maturity and yield_val:
                    if country not in benchmark_data:
                        benchmark_data[country] = {}
                    benchmark_data[country][maturity] = yield_val

        self.data['sections']['benchmark_yields'] = benchmark_data

    def generate_summary(self):
        """Generate summary statistics for the home page"""
        print("Generating Summary Statistics...")

        summary = {
            'total_countries': len(set(
                list(self.data['sections'].get('sovereign_yields', {}).get('domestic_currency', {}).get('1Y', {}).keys()) +
                list(self.data['sections'].get('sovereign_yields', {}).get('usd_denominated', {}).keys())
            )),
            'fx_pairs': len(self.data['sections'].get('fx_rates', {})),
            'credit_ratings': len(self.data['sections'].get('credit_ratings', {})),
            'corporate_ratings': len(self.data['sections'].get('corporate_yields', {})),
            'key_metrics': {}
        }

        # Add key metrics from each section
        if 'sovereign_yields' in self.data['sections']:
            usa_10y = self.data['sections']['sovereign_yields'].get('domestic_currency', {}).get('10Y', {}).get('USA')
            if usa_10y:
                summary['key_metrics']['USA_10Y_Treasury'] = f"{usa_10y:.3f}%"

        if 'corporate_yields' in self.data['sections']:
            aaa_yield = self.data['sections']['corporate_yields'].get('AAA', {}).get('effective_yield')
            if aaa_yield:
                summary['key_metrics']['USA_AAA_Corporate'] = f"{aaa_yield:.3f}%"

        self.data['sections']['summary'] = summary

    def save_json(self, output_path):
        """Save extracted data to JSON file"""
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, indent=2, default=str, ensure_ascii=False)

            print(f"Saved dashboard data to: {output_path}")
            print(f"  File size: {os.path.getsize(output_path)} bytes")
            print(f"  Sections: {list(self.data['sections'].keys())}")

        except Exception as e:
            raise RuntimeError(f"Failed to save JSON: {e}")

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()

    def run(self, output_path):
        """Main ETL process"""
        try:
            print("Starting Markets Dashboard ETL Process...")
            print("=" * 60)

            self.load_workbook()

            # Extract all data sections
            self.extract_sovereign_yields()
            self.extract_corporate_yields()
            # FX rates now fetched from ExchangeRate-API - see backend/src/api/fx_rates.py
            self.data['sections']['fx_rates'] = {}  # Placeholder (deprecated)
            self.extract_central_bank_rates()
            self.extract_credit_ratings()
            self.extract_benchmark_yields()
            self.generate_summary()

            self.save_json(output_path)

            print("=" * 60)
            print("ETL Process completed successfully!")

            return True

        except Exception as e:
            print(f"ETL Process failed: {e}")
            return False
        finally:
            self.close()


def main():
    """Main entry point"""
    # Get paths
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    excel_file = project_root / "backend" / "data" / "excel" / "Markets Dashboard (Macro Enabled) (version 3).xlsm"
    output_file = project_root / "backend" / "storage" / "dashboard.json"

    if not excel_file.exists():
        print(f"Excel file not found: {excel_file}")
        sys.exit(1)

    # Run ETL
    etl = DashboardETL(str(excel_file))
    success = etl.run(str(output_file))

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
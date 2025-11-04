#!/usr/bin/env python3
"""
ETL script to extract USA Historical Yields from Markets Dashboard Excel.
Reads the USA sheet and extracts 3-month historical yield data for all maturities.
"""

import json
import os
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl


def extract_usa_historical_yields(excel_file_path, output_file_path):
    """
    Extract historical yield data from USA sheet

    Args:
        excel_file_path: Path to Markets Dashboard Excel file
        output_file_path: Path to output JSON file
    """
    try:
        # Load workbook (not read-only for faster formula evaluation)
        print(f"Loading workbook: {excel_file_path}")
        wb = openpyxl.load_workbook(excel_file_path, read_only=False, data_only=True, keep_vba=False)

        if 'USA' not in wb.sheetnames:
            raise ValueError("USA sheet not found in workbook")

        ws = wb['USA']
        print("USA sheet loaded successfully")

        # Extract dates from row 31 (header row) - ONLY dates with actual yield data
        print("Extracting date headers with actual data...")
        dates = []
        date_cols = []

        # Start from column D (column 4) and only include dates that have non-null yield data
        for col in range(4, 200):
            date_val = ws.cell(31, col).value
            # Check if this column has actual yield data (using 10Y as test)
            yield_test = ws.cell(42, col).value  # 10Y row

            if date_val and isinstance(date_val, datetime) and yield_test is not None:
                dates.append(date_val)
                date_cols.append(col)

        if not dates:
            raise ValueError("No date headers with actual data found in row 31")

        print(f"Found {len(dates)} date columns with actual data from {dates[0]} to {dates[-1]}")

        # Calculate 3-month cutoff from the LAST ACTUAL date (not future dates)
        last_date = dates[-1]
        three_months_ago = last_date - timedelta(days=90)

        # Find the starting column index for 3-month window
        start_idx = 0
        for i, date in enumerate(dates):
            if date >= three_months_ago:
                start_idx = i
                break

        # Filter to last 3 months
        filtered_dates = dates[start_idx:]
        filtered_cols = date_cols[start_idx:]

        print(f"Using {len(filtered_dates)} days from {filtered_dates[0]} to {filtered_dates[-1]}")

        # Define maturities and their row numbers (rows 32-44)
        maturities = [
            ('1M', 32),
            ('2M', 33),
            ('3M', 34),
            ('4M', 35),
            ('6M', 36),
            ('1Y', 37),
            ('2Y', 38),
            ('3Y', 39),
            ('5Y', 40),
            ('7Y', 41),
            ('10Y', 42),
            ('20Y', 43),
            ('30Y', 44)
        ]

        # Extract yield data for each maturity
        print("Extracting yield data by maturity...")
        yield_data = {}

        for maturity_label, row_num in maturities:
            yields = []
            for col in filtered_cols:
                cell_value = ws.cell(row_num, col).value
                # Convert to float and handle None/errors
                if cell_value is not None:
                    try:
                        yields.append(float(cell_value) * 100)  # Convert to percentage
                    except (ValueError, TypeError):
                        yields.append(None)
                else:
                    yields.append(None)

            yield_data[maturity_label] = yields
            print(f"  {maturity_label}: {len(yields)} data points")

        # Build output structure
        output = {
            'metadata': {
                'last_updated': datetime.now().isoformat(),
                'source_file': os.path.basename(excel_file_path),
                'generated_by': 'extract_usa_historical.py',
                'date_range': {
                    'start': filtered_dates[0].strftime('%Y-%m-%d'),
                    'end': filtered_dates[-1].strftime('%Y-%m-%d'),
                    'days': len(filtered_dates)
                }
            },
            'dates': [d.strftime('%Y-%m-%d') for d in filtered_dates],
            'maturities': yield_data
        }

        # Write to JSON file
        output_path = Path(output_file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w') as f:
            json.dump(output, f, indent=2)

        print(f"\nSuccessfully wrote historical yields to: {output_file_path}")
        print(f"Total maturities: {len(maturities)}")
        print(f"Date range: {len(filtered_dates)} days")

        return output

    except Exception as e:
        print(f"Error extracting USA historical yields: {e}")
        raise


if __name__ == '__main__':
    # Set paths
    project_root = Path(__file__).parent.parent
    excel_file = project_root / 'backend' / 'data' / 'excel' / 'Markets Dashboard (Macro Enabled) (version 3).xlsm'
    output_file = project_root / 'backend' / 'storage' / 'usa_historical_yields.json'

    # Run extraction
    extract_usa_historical_yields(str(excel_file), str(output_file))
